#!/usr/bin/env python3

"""
Ejemplo de cómo retomar el scraping desde donde se quedó si el programa falla.
Se basa en guardar el índice procesado en un archivo JSON (checkpoint.json).
Si el programa se detiene abruptamente, al relanzarlo, reanudará en el índice
siguiente al último procesado, a menos que se pase un índice manual por argumento.
"""

import argparse
import logging
import time
import os
import random
import json
import pandas as pd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
from webdriver_manager.chrome import ChromeDriverManager

from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    ElementClickInterceptedException,
    ElementNotInteractableException,
    StaleElementReferenceException
)
from openpyxl import load_workbook
import shutil
import datetime

def backup_excel_file(excel_path):
    """Crea una copia de seguridad del archivo Excel antes de modificarlo."""
    backup_path = excel_path.replace(".xlsx", f"_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    shutil.copy2(excel_path, backup_path)
    print(f"Copia de seguridad creada: {backup_path}")


def save_row_to_excel(excel_path, data, index):
    """Guarda una fila específica en el archivo Excel sin sobrescribir todo el archivo."""
    temp_path = excel_path.replace(".xlsx", "_temp.xlsx")
    try:
        wb = load_workbook(excel_path)
        ws = wb.active

        row_values = data.loc[index].tolist()
        for col_num, value in enumerate(row_values, start=1):
            ws.cell(row=index + 2, column=col_num, value=value)

        wb.save(temp_path)
        wb.close()

        # Renombrar solo si todo fue exitoso
        os.replace(temp_path, excel_path)
        if debug_flag:
            print(f"✅ Guardado seguro en {excel_path}")

    except Exception as e:
        print(f"⚠️ Error en guardado atómico: {e}")


debug_flag = False

# ---------- ids --------------

registro_input_id = "ctl00_ContentPlaceHolder1_txtNumeroRegistro"
buscar_button_id = "ctl00_ContentPlaceHolder1_btnBuscar"
lupa_button_xpath = "//input[@id='ctl00_ContentPlaceHolder1_gvDatosBusqueda_ctl02_btnFicha']"
drop_down_estado_vigencia = "ctl00_ContentPlaceHolder1_ddlEstado"

# -----------------------------

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 14_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Mobile/15A5370a Safari/604.1"
]

# Columnas utilizadas
required_columns = [
    "Nombre", "Referencia_de_tramite", "Equivalencia_Terapeutica_o_Biosimilar", "Titular",
    "Estado_del_Registro", "Resolucion_Inscribase", "Fecha_Inscribase", "Ultima_Renovacion",
    "Fecha_Proxima_Renovacion", "Regimen", "Via_Administracion", "Condicion_de_Venta",
    "Expende_tipo_establecimiento", "Indicacion", "Nombre_PA", "Concentracion", "Unidad_medida"
]

CHECKPOINT_FILE = "checkpoint.json"

# --------------- versioning -----


VERSION = "0214"


# -------------- logging --------------------------------

def welcoming():
    ascii_art = r"""
    .___  ___________________   ____  __.___.____    .____     _____________________ 
    |   |/   _____/\______   \ |    |/ _|   |    |   |    |    \_   _____/\______   \
    |   |\_____  \  |     ___/ |      < |   |    |   |    |     |    __)_  |       _/
    |   |/        \ |    |     |    |  \|   |    |___|    |___  |        \ |    |   \
    |___/_______  / |____|     |____|__ \___|_______ \_______ \/_______  / |____|_  /
                \/                     \/           \/       \/        \/         \/ 


                RecetAI - ISP Killer
    (Programa para obtener información de medicamentos)
    """

    print(ascii_art)
    print("\n¡Bienvenido al script ISP Killer de RecetAI!\n")
    print(f"Version 0.{VERSION}")




# -------------- FUNCIONES CHECKPOINT -------------------

def load_checkpoint():
    """Lee el índice desde un archivo JSON si existe. Retorna 0 si no existe."""
    if os.path.exists(CHECKPOINT_FILE):
        try:
            with open(CHECKPOINT_FILE, "r") as f:
                data = json.load(f)
            return data.get("last_index", 0)
        except:  # si falla la lectura
            return 0
    else:
        return 0


def save_checkpoint(index):
    """Guarda el índice (fila) en un archivo JSON para retomar luego."""
    data = {"last_index": index}
    with open(CHECKPOINT_FILE, "w") as f:
        json.dump(data, f)

# -------------- FIN FUNCIONES CHECKPOINT ---------------

def simulate_typing(element, text):
    for char in text:
        element.send_keys(char)
        time.sleep(random.uniform(0.2, 0.3))


def select_estado(driver, wait, logger, estado, wait_multiplier=1):
    dropdown_element = wait.until(
        EC.element_to_be_clickable((By.ID, drop_down_estado_vigencia))
    )
    select_estado = Select(dropdown_element)
    select_estado.select_by_value(estado)
    time.sleep(0.1 / wait_multiplier)
    logger.debug(f"Estado de vigencia seleccionado: {estado}")


def ingresar_registro(driver, wait, logger, registro, wait_multiplier=1):
    registro_input = wait.until(
        EC.presence_of_element_located((By.ID, registro_input_id))
    )
    registro_input.clear()
    simulate_typing(registro_input, str(registro))
    wait.until(EC.element_to_be_clickable((By.ID, buscar_button_id))).click()
    logger.debug(f"Registro ingresado: {registro}")


def obtener_cantidad_resultados(driver, wait, logger, registro, wait_multiplier=1, max_retries=5) -> int:
    """
    Intenta obtener la cantidad de resultados encontrados en la búsqueda.
    Si no se puede leer el elemento, reintenta varias veces antes de fallar.
    """
    for intento in range(1, max_retries + 1):
        try:
            span_cantidad = wait.until(
                EC.visibility_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_lblCantidadEC"))
            )
            cantidad_str = span_cantidad.text.strip() or "0"
            return int(cantidad_str)
        except Exception as e:
            logger.warning(f"[{registro}] Intento {intento}/{max_retries} - No se pudo leer lblCantidadEC: {e}")
            time.sleep(2 / wait_multiplier)  # Esperar antes de reintentar
    
    logger.error(f"[{registro}] No se pudo obtener lblCantidadEC tras {max_retries} intentos. Saltando registro.")
    return 0  # Retornar 0 para que el script no se caiga



def intentar_click_lupa(driver, wait, logger, registro, wait_multiplier=1, intentos=5) -> bool:
    for intento in range(1, intentos + 1):
        try:
            time.sleep(0.2 / wait_multiplier)
            lupa_elem = wait.until(EC.element_to_be_clickable((By.XPATH, lupa_button_xpath)))
            lupa_elem.click()
            logger.debug(f"[{registro}] Clic en la lupa (intento {intento}/{intentos}).")
            return True
        except Exception as e:
            logger.debug(f"[{registro}] Falló clic lupa (intento {intento}/{intentos}): {e}")
            time.sleep(0.5 / wait_multiplier)
    logger.warning(f"[{registro}] No se pudo hacer clic en la lupa tras {intentos} intentos.")
    return False


# ---------------- FUNCIONES PARA EXTRAER DATOS --------------

def safe_extract(registro, driver, by, identifier, wait_multiplier=1, logger=None):
    try:
        elem = driver.find_element(by, identifier)
        ActionChains(driver).move_to_element(elem).perform()
        time.sleep(0.3 / wait_multiplier)
        return elem.text.strip()
    except NoSuchElementException:
        if logger:
            logger.warning(f"[{registro}] No se encontró el elemento {identifier}.")
        return "N/A"
    except Exception as e:
        if logger:
            logger.warning(f"[{registro}] Error al extraer {identifier}: {e}")
        return "N/A"


def extract_data(driver, registro, wait_multiplier=1, logger=None):
    def se(by, identifier):
        return safe_extract(
            registro=registro,
            driver=driver,
            by=by,
            identifier=identifier,
            wait_multiplier=wait_multiplier,
            logger=logger
        )
    extracted_data = {
        "Nombre": se(By.ID, "ctl00_ContentPlaceHolder1_lblNombre"),
        "Referencia_de_tramite": se(By.ID, "ctl00_ContentPlaceHolder1_lblRefTramite"),
        "Equivalencia_Terapeutica_o_Biosimilar": se(By.ID, "ctl00_ContentPlaceHolder1_lblEquivalencia"),
        "Titular": se(By.ID, "ctl00_ContentPlaceHolder1_lblEmpresa"),
        "Estado_del_Registro": se(By.ID, "ctl00_ContentPlaceHolder1_lblEstado"),
        "Resolucion_Inscribase": se(By.ID, "ctl00_ContentPlaceHolder1_lblResInscribase"),
        "Fecha_Inscribase": se(By.ID, "ctl00_ContentPlaceHolder1_lblFchInscribase"),
        "Ultima_Renovacion": se(By.ID, "ctl00_ContentPlaceHolder1_lblUltimaRenovacion"),
        "Fecha_Proxima_Renovacion": se(By.ID, "ctl00_ContentPlaceHolder1_lblProxRenovacion"),
        "Regimen": se(By.ID, "ctl00_ContentPlaceHolder1_lblRegimen"),
        "Via_Administracion": se(By.ID, "ctl00_ContentPlaceHolder1_lblViaAdministracion"),
        "Condicion_de_Venta": se(By.ID, "ctl00_ContentPlaceHolder1_lblCondicionVenta"),
        "Expende_tipo_establecimiento": se(By.ID, "ctl00_ContentPlaceHolder1_lblExTipoEstablecimiento"),
        "Indicacion": se(By.ID, "ctl00_ContentPlaceHolder1_lblIndicacion"),
        "Nombre_PA": se(By.ID, "ctl00_ContentPlaceHolder1_gvFormulas_ctl02_lblNombreElemento"),
        "Concentracion": se(By.ID, "ctl00_ContentPlaceHolder1_gvFormulas_ctl02_lblConcentracion"),
        "Unidad_medida": se(By.ID, "ctl00_ContentPlaceHolder1_gvFormulas_ctl02_lblUnidadMedida")
    }
    return extracted_data


# ----------------- PROCESO DE REGISTROS ------------------

def process_single_record(
    driver, wait, data, index, row, logger, wait_multiplier=1
    ) -> bool:
    registro = row["Registro"]
    estados_a_probar = ["Sí", "No", "Suspendido"]

    for estado in estados_a_probar:
        select_estado(driver, wait, logger, estado, wait_multiplier)
        ingresar_registro(driver, wait, logger, registro, wait_multiplier)

        cantidad_encontrada = obtener_cantidad_resultados(
            driver, wait, logger, registro, wait_multiplier
        )
        if cantidad_encontrada > 0:
            if not intentar_click_lupa(driver, wait, logger, registro, wait_multiplier):
                continue  # intenta siguiente estado

            extracted_data = extract_data(
                driver=driver,
                registro=registro,
                wait_multiplier=wait_multiplier,
                logger=logger
            )
            for key, value in extracted_data.items():

                data.at[index, key] = str(value)
                if debug_flag:
                    print(data.at[index, key])
 

            driver.back()
            wait.until(EC.presence_of_element_located((By.ID, registro_input_id)))
            return True
        else:
            logger.debug(f"[{registro}] Con estado '{estado}' se encontraron 0 registros.")
    return False  # no se encontró info en ninguno de los estados

def process_records(
    driver, wait, data, start_index, end_index,
    temp_excel_path, excel_path, logger, total_registros, wait_multiplier=1,setup_driver=None
    ):
    checkpoint_index = load_checkpoint()
    if start_index == -1:
        logger.info(f"Checkpoint indica que ya se procesaron registros hasta la fila {checkpoint_index}. "
                    "Reanudando desde allí...")
        start_index = checkpoint_index

    procesados = 0
    errores_consecutivos = 0  # Contador de errores seguidos

    for index in range(start_index, end_index):
        row = data.iloc[index]
        registro = row["Registro"]
        logger.info(f"Procesando registro {registro} (Fila {index + 1} de {total_registros})")

        try:
            se_encontro_info = process_single_record(
                driver=driver,
                wait=wait,
                data=data,
                index=index,
                row=row,
                logger=logger,
                wait_multiplier=wait_multiplier
            )

            if not se_encontro_info:
                logger.warning(f"[{registro}] No se encontraron registros para ninguno de los estados.")
                errores_consecutivos += 1  # Contamos fallos consecutivos
            else:
                errores_consecutivos = 0  # Reset al encontrar un registro válido

            procesados += 1
            logger.info(f"Registros procesados: {procesados} de {total_registros}")

            # Guardamos el checkpoint
            save_checkpoint(index + 1)

            save_row_to_excel(excel_path, data, index)


            # Si hay demasiados errores consecutivos, reiniciar navegador
            if errores_consecutivos >= 10:
                logger.error(f"Se han detectado {errores_consecutivos} errores seguidos. Reiniciando navegador...")
                driver.quit()
                driver = setup_driver()  # Reiniciamos el driver
                wait = WebDriverWait(driver, 5)
                errores_consecutivos = 0  # Reset contador

        except Exception as e:
            logger.error(f"Error procesando fila {index}: {e}")
            logger.info("Deteniendo el proceso para reanudar en la siguiente ejecución.")
            break  # rompe el bucle, en la siguiente ejecución reanudará donde se quedó



# -------------------------- MAIN -----------------------------

def main(excel_path, start_index=None, wait_multiplier=1):


    logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")
    logger = logging.getLogger()
    
    # genera un backup
    backup_excel_file(excel_path)


    def setup_driver():
        opts = Options()
        opts.add_argument("--incognito")
        opts.add_argument("--start-maximized")
        opts.add_argument("--headless")
        opts.add_argument("--disable-gpu")
        opts.add_argument(f"user-agent={random.choice(USER_AGENTS)}")
        return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)

    driver = None
    try:
        logger.info("=== Iniciando Selenium ===")
        driver = setup_driver()
        wait = WebDriverWait(driver, 5)

        temp_excel_path = excel_path.replace(".xlsx", "_temp.xlsx")
        logger.info(f"Cargando archivo Excel desde: {excel_path}")
        data = pd.read_excel(excel_path, dtype=str)

        # Aseguramos columnas requeridas
        for col in required_columns:
            if col not in data.columns:
                data[col] = "N/A"

        url = "https://registrosanitario.ispch.gob.cl/"
        total_registros = len(data)

        driver.get(url)
        logger.info(f"Navegando a {url}")

        # Marcar el checkbox si es necesario
        try:
            checkbox_id = "ctl00_ContentPlaceHolder1_chkTipoBusqueda_3"
            checkbox = wait.until(EC.element_to_be_clickable((By.ID, checkbox_id)))
            driver.execute_script("arguments[0].scrollIntoView();", checkbox)
            time.sleep(random.uniform(1, 2) / wait_multiplier)
            if not checkbox.is_selected():
                checkbox.click()
                time.sleep(random.uniform(1, 2) / wait_multiplier)
            checkbox = driver.find_element(By.ID, checkbox_id)
            if not checkbox.is_selected():
                logger.error("El checkbox no quedó marcado correctamente.")
        except Exception as e:
            logger.error(f"Error al interactuar con el checkbox: {e}")

        # Si el usuario no pasa start_index, usamos el checkpoint
        if start_index is None:
            logger.info("No se proporcionó --index, usando checkpoint si existe...")
            start_index = load_checkpoint()
        else:
            logger.info(f"Se usará start_index={start_index}")

        process_records(
            driver=driver,
            wait=wait,
            data=data,
            start_index=start_index,
            end_index=len(data),
            temp_excel_path=temp_excel_path,
            excel_path=excel_path,
            logger=logger,
            total_registros=total_registros,
            wait_multiplier=wait_multiplier,
            setup_driver=setup_driver
        )

        logger.info("Proceso finalizado")
    except KeyboardInterrupt: # Ctrl + Z
        logger.info("Interrupción de teclado detectada. Finalizando.")
        if driver:
            logger.info("Cerrando el navegador.")
            driver.quit()
        return 2 # representa el cierre del progrmaa

    except Exception as e:
        logger.error(f"Ha ocurrido un error crítico: {e}")

    finally:
        if driver:
            logger.info("Cerrando el navegador.")
            driver.quit()
        logger.info("Programa finalizado.")

        return -1 # representa un posible error critico


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=f"Programa para procesar registros sanitarios con Selenium. Versión 0.{VERSION}")
    parser.add_argument(
        "--x",
        required=False,
        default="Listado_de_registros.xlsx",
        help="Ruta del archivo Excel (.xlsx) que se desea procesar."
    )
    parser.add_argument(
        "--index",
        required=False,
        default=-1,
        type=int,
        help="Fila de Excel desde la cual comenzar el procesamiento. Si se omite, retoma del checkpoint.json"
    )
    parser.add_argument(
        "--waitx",
        required=False,
        type=int,
        default=1,
        help="Valor: a mayor 'waitx', más lentos los tiempos de espera"
    )
    args = parser.parse_args()

    excel_path = args.x
    start_index = args.index  # puede ser None
    wait_multiplier = args.waitx

    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
    logger = logging.getLogger()

    welcoming()

    output = main(excel_path, start_index, wait_multiplier=wait_multiplier)

    if output == -1: # reiniciar el programa pero desde el json
        main(excel_path, start_index=-1, wait_multiplier=wait_multiplier)
    